import openpyxl

inv_file = openpyxl.load_workbook("FANG.xlsx")
product_list =  inv_file["Sheet1"]


# Name of company will be key and the product count will be the value
products_per_supplier = {} # Exercise 1
total_value_per_supplier = {}
products_under_10_inv = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        # current_num_products = products_per_supplier.get(supplier_name)
        # products_per_supplier[supplier_name] = current_num_products + 1
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        # current_total_value = total_value_per_supplier[supplier_name]
        # total_value_per_supplier[supplier_name] = current_total_value + inventory * price

        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)



    # Add value for total inventory price
    inventory_price.value = inventory * price




print(f"The total amount of products per supplier -> {products_per_supplier}")
print(f"The total value for each  supplier is : {total_value_per_supplier}")
print(products_under_10_inv)
inv_file .save("Inventory_with_total_value.xlsx")
